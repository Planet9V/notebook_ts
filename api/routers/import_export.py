"""Import / Export API router.

Handles CSV and XLSX file import (preview + execute) for customer/contact
bulk operations, and streaming export of customer/contact data.

Follows the same patterns as api/routers/customers.py and contacts.py:
  - repo_query / repo_create / repo_update for SurrealDB ops
  - Pydantic request/response models from api/models.py
  - loguru for logging
"""

import csv
import io
import json
import os
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from loguru import logger

from api.models import (
    ImportErrorDetail,
    ImportExecuteRequest,
    ImportOptions,
    ImportPreviewResponse,
    ImportResultResponse,
    ImportWarningDetail,
)
from open_notebook.database.repository import (
    repo_create,
    repo_insert,
    repo_query,
    repo_update,
)

router = APIRouter()


# ──────────────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────────────

CUSTOMER_TARGET_FIELDS = [
    "name", "website", "description", "industry",
    "street_address", "street_address_2", "city", "state", "postal_code", "country",
    "phone", "phone_alt", "fax", "email",
    "salesperson", "lead_source", "annual_revenue", "employee_count",
    "customer_type", "tier", "status",
    "linkedin_url", "twitter_url", "facebook_url",
    "tags", "internal_notes",
]

CONTACT_TARGET_FIELDS = [
    "contact_name", "first_name", "last_name",
    "contact_email", "contact_phone", "contact_mobile",
    "contact_title", "contact_department", "contact_seniority",
    "contact_linkedin_url",
]

# Fuzzy column header → target field mapping.
# Keys are lowercase variations; values are canonical field names.
_HEADER_ALIASES: Dict[str, str] = {
    # Customer name
    "company": "name",
    "company name": "name",
    "organization": "name",
    "org": "name",
    "account": "name",
    "account name": "name",
    "customer": "name",
    "customer name": "name",
    "name": "name",
    # Email
    "email": "email",
    "e-mail": "email",
    "email address": "email",
    "company email": "email",
    # Phone
    "phone": "phone",
    "phone number": "phone",
    "tel": "phone",
    "telephone": "phone",
    "main phone": "phone",
    # State / Province
    "state": "state",
    "province": "state",
    "region": "state",
    "state/province": "state",
    # City
    "city": "city",
    "town": "city",
    # Postal
    "zip": "postal_code",
    "zip code": "postal_code",
    "postal code": "postal_code",
    "postcode": "postal_code",
    # Country
    "country": "country",
    # Address
    "address": "street_address",
    "street": "street_address",
    "street address": "street_address",
    "address 2": "street_address_2",
    "suite": "street_address_2",
    # Website
    "website": "website",
    "url": "website",
    "web": "website",
    # Industry
    "industry": "industry",
    "sector": "industry",
    # Description
    "description": "description",
    "notes": "internal_notes",
    # Contact name
    "contact": "contact_name",
    "contact name": "contact_name",
    "full name": "contact_name",
    "contact person": "contact_name",
    "first name": "first_name",
    "last name": "last_name",
    # Contact email/phone
    "contact email": "contact_email",
    "contact phone": "contact_phone",
    "contact mobile": "contact_mobile",
    "mobile": "contact_mobile",
    # Title
    "title": "contact_title",
    "job title": "contact_title",
    "position": "contact_title",
    "contact title": "contact_title",
    # Department
    "department": "contact_department",
    "contact department": "contact_department",
    # LinkedIn
    "linkedin": "linkedin_url",
    "linkedin url": "linkedin_url",
    "contact linkedin": "contact_linkedin_url",
    # Revenue / Size
    "revenue": "annual_revenue",
    "annual revenue": "annual_revenue",
    "employees": "employee_count",
    "employee count": "employee_count",
    "size": "employee_count",
    # Sales
    "salesperson": "salesperson",
    "sales rep": "salesperson",
    "lead source": "lead_source",
    "source": "lead_source",
}


# ──────────────────────────────────────────────────────────────────────────────
# File parsing helpers
# ──────────────────────────────────────────────────────────────────────────────


def _parse_csv(raw_bytes: bytes) -> Tuple[List[str], List[List[str]]]:
    """Parse CSV bytes into (headers, data_rows). Handles BOM (utf-8-sig)."""
    text = raw_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    headers = [h.strip() for h in rows[0]]
    data = rows[1:]
    return headers, data


def _parse_xlsx(raw_bytes: bytes) -> Tuple[List[str], List[List[str]]]:
    """Parse XLSX bytes into (headers, data_rows) using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []

    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()

    if not rows:
        return [], []
    headers = [h.strip() for h in rows[0]]
    data = rows[1:]
    return headers, data


def _parse_file(filename: str, raw_bytes: bytes) -> Tuple[List[str], List[List[str]]]:
    """Dispatch to CSV or XLSX parser based on file extension."""
    ext = os.path.splitext(filename)[1].lower()
    if ext in (".xlsx", ".xls"):
        return _parse_xlsx(raw_bytes)
    # Default to CSV for .csv and anything else
    return _parse_csv(raw_bytes)


# ──────────────────────────────────────────────────────────────────────────────
# Auto-mapping
# ──────────────────────────────────────────────────────────────────────────────


def _auto_map_columns(headers: List[str]) -> Dict[str, str]:
    """Return {column_header: target_field} for headers that fuzzy-match known aliases."""
    mapping: Dict[str, str] = {}
    for header in headers:
        normalised = header.strip().lower()
        if normalised in _HEADER_ALIASES:
            mapping[header] = _HEADER_ALIASES[normalised]
    return mapping


# ──────────────────────────────────────────────────────────────────────────────
# Import endpoints
# ──────────────────────────────────────────────────────────────────────────────


@router.post("/customers/import/preview", response_model=ImportPreviewResponse)
async def import_preview(file: UploadFile = File(...)):
    """Parse an uploaded CSV/XLSX file and return a preview with auto-mapping suggestions."""
    try:
        raw = await file.read()
        if len(raw) > 10 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="File too large. Maximum size is 10MB.")
        filename = file.filename or "upload.csv"
        headers, data_rows = _parse_file(filename, raw)

        if not headers:
            raise HTTPException(status_code=400, detail="File contains no headers")

        suggested_mapping = _auto_map_columns(headers)
        sample_rows = data_rows[:5]

        return ImportPreviewResponse(
            file_name=filename,
            total_rows=len(data_rows),
            columns=headers,
            sample_rows=sample_rows,
            suggested_mapping=suggested_mapping,
            available_customer_fields=CUSTOMER_TARGET_FIELDS,
            available_contact_fields=CONTACT_TARGET_FIELDS,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error previewing import file: {e}")
        raise HTTPException(status_code=500, detail=f"Error previewing import file: {e}")


@router.post("/customers/import", response_model=ImportResultResponse)
async def import_execute(
    file: UploadFile = File(...),
    import_config: str = Form(...),
):
    """Execute a customer/contact import from a CSV/XLSX file.

    Accepts multipart/form-data with:
      - file: the CSV or XLSX upload
      - import_config: JSON string matching ImportExecuteRequest schema
    """
    try:
        config = ImportExecuteRequest.model_validate_json(import_config)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Invalid import_config JSON: {e}")

    batch_id = f"import_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
    mapping = config.column_mapping
    options = config.options

    # Read and parse file
    try:
        raw = await file.read()
        if len(raw) > 10 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="File too large. Maximum size is 10MB.")
        filename = file.filename or config.file_name
        headers, data_rows = _parse_file(filename, raw)
    except Exception as e:
        logger.error(f"Error parsing import file: {e}")
        raise HTTPException(status_code=400, detail=f"Error parsing import file: {e}")

    if not headers or not data_rows:
        raise HTTPException(status_code=400, detail="File has no data rows")

    # Build column index lookup
    col_index: Dict[str, int] = {h: i for i, h in enumerate(headers)}

    # Separate customer vs contact field mappings
    contact_field_prefixes = {
        "contact_name", "first_name", "last_name",
        "contact_email", "contact_phone", "contact_mobile",
        "contact_title", "contact_department", "contact_seniority",
        "contact_linkedin_url",
    }

    customer_mapping: Dict[str, str] = {}  # col_header → customer field
    contact_mapping: Dict[str, str] = {}   # col_header → contact field
    for col_header, target_field in mapping.items():
        if target_field in contact_field_prefixes:
            contact_mapping[col_header] = target_field
        else:
            customer_mapping[col_header] = target_field

    has_contact_fields = bool(contact_mapping)

    # Pre-fetch existing customer names for duplicate detection
    existing_customers: Dict[str, dict] = {}
    if options.duplicate_strategy in ("skip", "update"):
        try:
            all_customers = await repo_query("SELECT * FROM customer;")
            for c in all_customers:
                cname = str(c.get("name", "")).strip().lower()
                if cname:
                    existing_customers[cname] = c
        except Exception as e:
            logger.warning(f"Could not fetch existing customers for dedup: {e}")

    # Counters
    customers_created = 0
    customers_updated = 0
    customers_skipped = 0
    contacts_created = 0
    notebooks_created = 0
    errors: List[ImportErrorDetail] = []
    warnings: List[ImportWarningDetail] = []

    # Batch buffers for contacts and notebooks (flush after loop)
    pending_contacts: List[Dict[str, Any]] = []
    pending_notebooks: List[Dict[str, Any]] = []
    contact_row_map: List[int] = []  # Track row indices for error reporting
    notebook_row_map: List[int] = []

    for row_idx, row in enumerate(data_rows, start=2):  # row 1 = header
        # ── Build customer dict ──
        customer_data: Dict[str, Any] = {}
        for col_header, field in customer_mapping.items():
            idx = col_index.get(col_header)
            if idx is not None and idx < len(row):
                value = row[idx].strip()
                if value:
                    # Coerce numeric fields
                    if field == "annual_revenue":
                        try:
                            customer_data[field] = float(value.replace(",", "").replace("$", ""))
                        except ValueError:
                            warnings.append(ImportWarningDetail(
                                row=row_idx, field=field,
                                warning=f"Could not parse '{value}' as number",
                            ))
                    elif field == "employee_count":
                        try:
                            customer_data[field] = int(value.replace(",", ""))
                        except ValueError:
                            warnings.append(ImportWarningDetail(
                                row=row_idx, field=field,
                                warning=f"Could not parse '{value}' as integer",
                            ))
                    else:
                        customer_data[field] = value

        # Name is required
        cust_name = customer_data.get("name", "").strip()
        if not cust_name:
            errors.append(ImportErrorDetail(
                row=row_idx, field="name",
                error="Missing required customer name",
                data=customer_data,
            ))
            continue

        # Apply defaults
        customer_data.setdefault("customer_type", options.default_customer_type)
        customer_data.setdefault("tier", options.default_tier)
        customer_data.setdefault("lead_source", options.default_lead_source)
        customer_data["import_batch_id"] = batch_id
        customer_data["import_source"] = filename
        if options.tags:
            existing_tags = customer_data.get("tags", [])
            if isinstance(existing_tags, list):
                customer_data["tags"] = list(set(existing_tags + options.tags))
            else:
                customer_data["tags"] = list(options.tags)

        # ── Duplicate detection ──
        name_key = cust_name.lower()
        customer_record: Optional[dict] = None

        if name_key in existing_customers:
            existing = existing_customers[name_key]
            if options.duplicate_strategy == "skip":
                customers_skipped += 1
                customer_record = existing
                warnings.append(ImportWarningDetail(
                    row=row_idx, field="name",
                    warning=f"Duplicate customer '{cust_name}' — skipped",
                ))
            elif options.duplicate_strategy == "update":
                try:
                    existing_id = str(existing.get("id", ""))
                    result = await repo_update("customer", existing_id, customer_data)
                    customer_record = result[0] if result else existing
                    customers_updated += 1
                except Exception as e:
                    errors.append(ImportErrorDetail(
                        row=row_idx, field="name",
                        error=f"Failed to update existing customer: {e}",
                        data=customer_data,
                    ))
                    continue
            else:
                # duplicate_strategy == "create" — fall through
                pass

        # Create new customer if not handled above
        if customer_record is None:
            try:
                result = await repo_create("customer", customer_data)
                if isinstance(result, list):
                    result = result[0] if result else {}
                customer_record = result
                customers_created += 1
                # Add to dedup cache
                existing_customers[name_key] = result
            except Exception as e:
                errors.append(ImportErrorDetail(
                    row=row_idx, field="name",
                    error=f"Failed to create customer: {e}",
                    data=customer_data,
                ))
                continue

        customer_id = str(customer_record.get("id", ""))

        # ── Create contact if contact fields mapped ──
        if has_contact_fields:
            contact_data: Dict[str, Any] = {"customer_id": customer_id, "source": "import"}
            for col_header, field in contact_mapping.items():
                idx = col_index.get(col_header)
                if idx is not None and idx < len(row):
                    value = row[idx].strip()
                    if value:
                        # Map prefixed contact fields to contact table fields
                        if field == "contact_name":
                            parts = value.split(None, 1)
                            contact_data["first_name"] = parts[0]
                            contact_data["last_name"] = parts[1] if len(parts) > 1 else ""
                        elif field == "contact_email":
                            contact_data["email"] = value
                        elif field == "contact_phone":
                            contact_data["phone"] = value
                        elif field == "contact_mobile":
                            contact_data["mobile"] = value
                        elif field == "contact_title":
                            contact_data["title"] = value
                        elif field == "contact_department":
                            contact_data["department"] = value
                        elif field == "contact_seniority":
                            contact_data["seniority"] = value
                        elif field == "contact_linkedin_url":
                            contact_data["linkedin_url"] = value
                        else:
                            contact_data[field] = value

            # Only create contact if we have at least a name
            if contact_data.get("first_name"):
                contact_data["import_batch_id"] = batch_id
                pending_contacts.append(contact_data)
                contact_row_map.append(row_idx)

        # ── Queue notebook if requested ──
        if options.create_notebooks and customer_id:
            notebook_data = {
                "name": cust_name,
                "customer_id": customer_id,
                "stage": options.notebook_stage,
                "import_batch_id": batch_id,
            }
            pending_notebooks.append(notebook_data)
            notebook_row_map.append(row_idx)

    # ── Flush batched contacts ──
    if pending_contacts:
        try:
            await repo_insert("contact", pending_contacts)
            contacts_created = len(pending_contacts)
        except Exception as e:
            # Fallback: try individual creates to identify which rows failed
            logger.warning(f"Batch contact insert failed, falling back to individual: {e}")
            for i, cd in enumerate(pending_contacts):
                try:
                    await repo_create("contact", cd)
                    contacts_created += 1
                except Exception as ie:
                    warnings.append(ImportWarningDetail(
                        row=contact_row_map[i], field="contact_name",
                        warning=f"Customer created but contact failed: {ie}",
                    ))

    # ── Flush batched notebooks ──
    if pending_notebooks:
        try:
            await repo_insert("notebook", pending_notebooks)
            notebooks_created = len(pending_notebooks)
        except Exception as e:
            # Fallback: try individual creates
            logger.warning(f"Batch notebook insert failed, falling back to individual: {e}")
            for i, nd in enumerate(pending_notebooks):
                try:
                    await repo_create("notebook", nd)
                    notebooks_created += 1
                except Exception as ie:
                    warnings.append(ImportWarningDetail(
                        row=notebook_row_map[i], field="notebook",
                        warning=f"Customer created but notebook failed: {ie}",
                    ))

    return ImportResultResponse(
        batch_id=batch_id,
        total_rows=len(data_rows),
        customers_created=customers_created,
        customers_updated=customers_updated,
        customers_skipped=customers_skipped,
        contacts_created=contacts_created,
        notebooks_created=notebooks_created,
        errors=errors,
        warnings=warnings,
    )


# ──────────────────────────────────────────────────────────────────────────────
# Export endpoints
# ──────────────────────────────────────────────────────────────────────────────


def _build_csv_stream(headers: List[str], rows: List[Dict[str, Any]]) -> io.StringIO:
    """Build an in-memory CSV from a list of dicts."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({k: row.get(k, "") for k in headers})
    buf.seek(0)
    return buf


def _build_xlsx_bytes(headers: List[str], rows: List[Dict[str, Any]]) -> bytes:
    """Build an in-memory XLSX workbook from a list of dicts."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append([str(row.get(h, "") or "") for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/customers/export")
async def export_customers(
    format: str = Query("csv", description="Export format: csv or xlsx"),
    filters: Optional[str] = Query(None, description="JSON filter object (reserved for future use)"),
    include_contacts: bool = Query(False, description="Include related contacts as extra columns"),
):
    """Export customers as a downloadable CSV or XLSX file."""
    try:
        customers = await repo_query("SELECT * FROM customer ORDER BY name ASC;")

        customer_headers = [
            "id", "name", "website", "description", "industry",
            "street_address", "city", "state", "postal_code", "country",
            "phone", "email",
            "salesperson", "lead_source", "annual_revenue", "employee_count",
            "customer_type", "tier", "status",
            "linkedin_url", "tags",
            "created", "updated",
        ]

        # Flatten list fields
        export_rows: List[Dict[str, Any]] = []
        for c in customers:
            row = {h: c.get(h, "") for h in customer_headers}
            if isinstance(row.get("tags"), list):
                row["tags"] = "; ".join(row["tags"])
            export_rows.append(row)

        if include_contacts:
            contacts = await repo_query("SELECT * FROM contact ORDER BY last_name ASC;")
            contacts_by_customer: Dict[str, List[dict]] = {}
            for ct in contacts:
                cid = str(ct.get("customer_id", ""))
                contacts_by_customer.setdefault(cid, []).append(ct)

            # Add contact columns
            contact_cols = ["contact_name", "contact_email", "contact_phone", "contact_title"]
            customer_headers.extend(contact_cols)
            for row in export_rows:
                cid = str(row.get("id", ""))
                cid_clean = cid.split(":", 1)[-1] if ":" in cid else cid
                related = contacts_by_customer.get(cid, []) or contacts_by_customer.get(cid_clean, [])
                if related:
                    first = related[0]
                    row["contact_name"] = f"{first.get('first_name', '')} {first.get('last_name', '')}".strip()
                    row["contact_email"] = first.get("email", "")
                    row["contact_phone"] = first.get("phone", "")
                    row["contact_title"] = first.get("title", "")

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

        if format.lower() == "xlsx":
            xlsx_data = _build_xlsx_bytes(customer_headers, export_rows)
            return StreamingResponse(
                io.BytesIO(xlsx_data),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="customers_export_{timestamp}.xlsx"'},
            )
        else:
            csv_buf = _build_csv_stream(customer_headers, export_rows)
            return StreamingResponse(
                iter([csv_buf.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="customers_export_{timestamp}.csv"'},
            )
    except Exception as e:
        logger.error(f"Error exporting customers: {e}")
        raise HTTPException(status_code=500, detail=f"Error exporting customers: {e}")


@router.get("/contacts/export")
async def export_contacts(
    format: str = Query("csv", description="Export format: csv or xlsx"),
    filters: Optional[str] = Query(None, description="JSON filter object (reserved for future use)"),
):
    """Export contacts as a downloadable CSV or XLSX file."""
    try:
        contacts = await repo_query("SELECT * FROM contact ORDER BY last_name ASC;")

        contact_headers = [
            "id", "first_name", "last_name",
            "email", "phone", "mobile",
            "title", "department", "seniority",
            "linkedin_url", "customer_id",
            "status", "tags", "source",
            "created", "updated",
        ]

        export_rows: List[Dict[str, Any]] = []
        for ct in contacts:
            row = {h: ct.get(h, "") for h in contact_headers}
            if isinstance(row.get("tags"), list):
                row["tags"] = "; ".join(row["tags"])
            export_rows.append(row)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

        if format.lower() == "xlsx":
            xlsx_data = _build_xlsx_bytes(contact_headers, export_rows)
            return StreamingResponse(
                io.BytesIO(xlsx_data),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="contacts_export_{timestamp}.xlsx"'},
            )
        else:
            csv_buf = _build_csv_stream(contact_headers, export_rows)
            return StreamingResponse(
                iter([csv_buf.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="contacts_export_{timestamp}.csv"'},
            )
    except Exception as e:
        logger.error(f"Error exporting contacts: {e}")
        raise HTTPException(status_code=500, detail=f"Error exporting contacts: {e}")
