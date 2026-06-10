from datetime import datetime, timezone
from typing import List, Optional, Set

from fastapi import APIRouter, HTTPException, Query
from loguru import logger

from api.models import (
    AssessmentAnswerResponse,
    AssessmentAnswerUpdate,
    AssessmentCreate,
    AssessmentReportResponse,
    AssessmentReportStats,
    AssessmentResponse,
    AssessmentSessionCreate,
    AssessmentSessionResponse,
    CategoryCoverage,
    ComplianceSnapshot,
    FacilityRollup,
    FrameworkRollup,
    CustomerComplianceRollup,
    AssessmentSessionDiffResponse,
    AssessmentSessionDiffItem,
)
from open_notebook.database.repository import (
    ensure_record_id,
    repo_create,
    repo_query,
    repo_update,
)

router = APIRouter()

def clean_id(val: str) -> str:
    if not val:
        return ""
    return str(val).split(":", 1)[-1]


CSET_COMPONENT_MAPPING = {
    "activedirectory": ["Active Directory"],
    "applicationserver": ["Application Server"],
    "audioswitch": ["Audio Switch"],
    "buildingautomationmanagementsystems": ["Building Automation Management Systems"],
    "dcs": ["DCS"],
    "dnsserver": ["DNS Server"],
    "databaseserver": ["Database Server"],
    "dispatchconsole": ["Dispatch Console"],
    "electronicsecuritysystem": ["Electronic Security System"],
    "emergencymedicalservicecommunicationshardware": ["Emergency Medical Service Communications Hardware"],
    "ethernetbackhaul": ["Ethernet Backhaul"],
    "firewall": ["Firewall", "VPN", "Unidirectional Device", "Link Encryption", "IDS", "IPS"],
    "handheldwirelessdevice": ["Handheld Wireless Device"],
    "historian": ["Historian", "Database Server", "Application Server", "Imaging Server"],
    "ids": ["IDS", "Firewall"],
    "ipcamera": ["IP Camera"],
    "ipphone": ["IP Phone"],
    "ips": ["IPS", "Firewall"],
    "imagingmodalitiesandequipment": ["Imaging Modalities and Equipment"],
    "imagingserver": ["Imaging Server"],
    "linkencryption": ["Link Encryption"],
    "mailserver": ["Mail Server"],
    "modem": ["Modem"],
    "multiprotocollabelswitching": ["Multi Protocol Label Switching"],
    "networkprinter": ["Network Printer"],
    "networkscannerandcopier": ["Network Scanner And Copier"],
    "opticalringsystem": ["Optical Ring System"],
    "partner": ["Partner"],
    "physiologicalmonitoringsystem": ["Physiological Monitoring System"],
    "poweroverethernet": ["Power Over Ethernet"],
    "publickiosk": ["Public Kiosk"],
    "rfidtransmitter": ["RFID Transmitter"],
    "radiosite": ["Radio Site"],
    "realtimelocationsystem": ["Real Time Location System"],
    "relaypanel": ["Relay Panel"],
    "router": ["Router"],
    "safetyinstrumentedsystem": ["Safety Instrumented System"],
    "securityinformationandeventmanagementsystem": ["Security Information And Event Management System"],
    "subscriberradio": ["Subscriber Radio"],
    "switch": ["Switch", "VLAN Switch", "VLAN Router", "Router", "Ethernet Backhaul", "Optical Ring System", "Multi Protocol Label Switching"],
    "terminalserver": ["Terminal Server"],
    "unidirectionaldevice": ["Unidirectional Device"],
    "vlanrouter": ["VLAN Router"],
    "vlanswitch": ["VLAN Switch"],
    "vpn": ["VPN", "Firewall"],
    "webserver": ["Web Server"],
    "windowsupdateserver": ["Windows Update Server"],
    "wirelessmodem": ["Wireless Modem"],
    "wirelessnetwork": ["Wireless Network"],
    "rtu": ["RTU", "Relay Panel"],
    "plc": ["DCS"],
}


async def get_active_cset_prefixes_for_session(sess_id: str, fw_id: str) -> Optional[set]:
    """Resolves active CSET component prefixes for the Components framework session."""
    if str(fw_id) != "regulation:Components":
        return None

    active_prefixes = {"Comp"}  # general component questions are always active

    try:
        sess_rec = ensure_record_id(sess_id)
        sess_res = await repo_query("SELECT assessment_id FROM $id", {"id": sess_rec})
        if not sess_res:
            return active_prefixes

        assessment_id = sess_res[0].get("assessment_id")
        if not assessment_id:
            return active_prefixes

        assess_res = await repo_query("SELECT customer_id FROM $id", {"id": ensure_record_id(assessment_id)})
        if not assess_res:
            return active_prefixes

        customer_id = assess_res[0].get("customer_id")
        if not customer_id:
            return active_prefixes

        notebooks = await repo_query(
            "SELECT id FROM notebook WHERE customer_id != NONE AND type::string(customer_id) = type::string($cust_id) AND (archived = false OR archived = None)",
            {"cust_id": customer_id}
        )
        if not notebooks:
            return active_prefixes

        notebook_id = notebooks[0]["id"]

        assets = await repo_query(
            "SELECT type FROM asset WHERE notebook_id = $nb_id",
            {"nb_id": notebook_id}
        )

        for asset in assets:
            asset_type = asset.get("type", "")
            if not asset_type:
                continue
            normalized_type = asset_type.lower().replace("_", "").replace(" ", "")
            mapped_prefixes = CSET_COMPONENT_MAPPING.get(normalized_type, [])
            for p in mapped_prefixes:
                active_prefixes.add(p)
            # Add basic capitalized/upper forms as fallback
            active_prefixes.add(asset_type.upper())
            active_prefixes.add(asset_type.capitalize())
            active_prefixes.add(asset_type)
            
    except Exception as e:
        logger.error(f"Error resolving active CSET prefixes: {e}")

    return active_prefixes



@router.post("/assessments", response_model=AssessmentResponse)
async def create_assessment(data: AssessmentCreate):
    """Link a customer profile to a compliance framework assessment."""
    try:
        cust_id = data.customer_id
        if ":" not in cust_id:
            cust_id = f"customer:{cust_id}"
            
        fw_id = data.framework_id
        if ":" not in fw_id:
            fw_id = f"regulation:{fw_id}"
            
        # Check if customer exists
        cust_check = await repo_query("SELECT id FROM $id", {"id": ensure_record_id(cust_id)})
        if not cust_check:
            raise HTTPException(status_code=404, detail="Customer profile not found")
            
        # Check if regulation exists
        reg_check = await repo_query("SELECT id FROM $id", {"id": ensure_record_id(fw_id)})
        if not reg_check:
            raise HTTPException(status_code=404, detail="Compliance framework not found")
            
        # Check for existing link
        loc_id = data.location_id
        if loc_id and ":" not in loc_id:
            loc_id = f"location:{loc_id}"

        if loc_id:
            existing = await repo_query(
                "SELECT * FROM assessment WHERE type::string(customer_id) = type::string($cust_id) AND type::string(framework_id) = type::string($fw_id) AND type::string(location_id) = type::string($loc_id)",
                {"cust_id": cust_id, "fw_id": fw_id, "loc_id": loc_id}
            )
        else:
            existing = await repo_query(
                "SELECT * FROM assessment WHERE type::string(customer_id) = type::string($cust_id) AND type::string(framework_id) = type::string($fw_id) AND location_id = NONE",
                {"cust_id": cust_id, "fw_id": fw_id}
            )
        
        if existing:
            rec = existing[0]
        else:
            insert_data = {
                "customer_id": cust_id,
                "framework_id": fw_id,
                "created_at": datetime.now(timezone.utc).isoformat() + "Z"
            }
            if loc_id:
                insert_data["location_id"] = loc_id
            rec = await repo_create("assessment", insert_data)
            if isinstance(rec, list):
                rec = rec[0]
                
        # Emit activity for new assessments
        if not existing:
            from api.routers.activity_emitter import emit_activity
            fw_name = str(reg_check[0].get("name", fw_id)) if reg_check else fw_id
            await emit_activity(
                customer_id=cust_id,
                activity_type="assessment_started",
                description=f"Compliance assessment started: {fw_name}",
                metadata={"assessment_id": str(rec["id"]), "framework_id": str(fw_id)},
            )

        return AssessmentResponse(
            id=str(rec["id"]),
            customer_id=str(rec["customer_id"]),
            framework_id=str(rec["framework_id"]),
            created_at=str(rec.get("created_at") or ""),
            location_id=str(rec.get("location_id")) if rec.get("location_id") else None
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating compliance assessment: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/assessments", response_model=List[AssessmentResponse])
async def get_assessments(customer_id: str, location_id: Optional[str] = None):
    """Retrieve all compliance framework assessments active for a customer."""
    try:
        cust_id = customer_id
        if ":" not in cust_id:
            cust_id = f"customer:{cust_id}"
            
        if location_id:
            if location_id == 'none':
                results = await repo_query(
                    "SELECT * FROM assessment WHERE type::string(customer_id) = type::string($cust_id) AND location_id = NONE ORDER BY created_at DESC",
                    {"cust_id": cust_id}
                )
            else:
                loc_id = location_id if ":" in location_id else f"location:{location_id}"
                results = await repo_query(
                    "SELECT * FROM assessment WHERE type::string(customer_id) = type::string($cust_id) AND type::string(location_id) = type::string($loc_id) ORDER BY created_at DESC",
                    {"cust_id": cust_id, "loc_id": loc_id}
                )
        else:
            results = await repo_query(
                "SELECT * FROM assessment WHERE type::string(customer_id) = type::string($cust_id) ORDER BY created_at DESC",
                {"cust_id": cust_id}
            )
        
        return [
            AssessmentResponse(
                id=str(row["id"]),
                customer_id=str(row["customer_id"]),
                framework_id=str(row["framework_id"]),
                created_at=str(row.get("created_at") or ""),
                location_id=str(row.get("location_id")) if row.get("location_id") else None
            )
            for row in results
        ]
    except Exception as e:
        logger.error(f"Error fetching assessments: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/assessments/{assessment_id}/sessions", response_model=AssessmentSessionResponse)
async def create_session(assessment_id: str, data: AssessmentSessionCreate):
    """Start a new CSET audit session milestone, optionally copying from the last completed session."""
    try:
        assess_id = assessment_id
        if ":" not in assess_id:
            assess_id = f"assessment:{assessment_id}"
            
        # Check if assessment link exists
        assess = await repo_query("SELECT * FROM $id", {"id": ensure_record_id(assess_id)})
        if not assess:
            raise HTTPException(status_code=404, detail="Assessment link not found")
            
        assessment_record = assess[0]
        fw_id = assessment_record.get("framework_id")
        
        # Create the new session
        session_data = {
            "assessment_id": assess_id,
            "session_name": data.session_name,
            "status": "IN_PROGRESS",
            "version_lock": fw_id,
            "created_at": datetime.now(timezone.utc).isoformat() + "Z",
            "completed_at": None
        }
        
        new_sess = await repo_create("assessment_session", session_data)
        if isinstance(new_sess, list):
            new_sess = new_sess[0]
            
        new_sess_id = str(new_sess["id"])
        
        # Perform carry-forward if requested and a prior session exists
        if data.carry_forward_prior:
            prior = await repo_query(
                "SELECT id, created_at FROM assessment_session WHERE type::string(assessment_id) = type::string($assess_id) ORDER BY created_at DESC LIMIT 2",
                {"assess_id": assess_id}
            )
            # Since LIMIT 2 returns the newly created one as index 0, check if a prior index 1 exists
            if len(prior) > 1:
                prior_sess_id = str(prior[1]["id"])
                logger.info(f"Carrying forward baseline answers from prior session {prior_sess_id} to new session {new_sess_id}")
                
                # Retrieve all answers from prior session
                prior_answers = await repo_query(
                    "SELECT * FROM assessment_answer WHERE type::string(session_id) = type::string($prior_sess_id)",
                    {"prior_sess_id": prior_sess_id}
                )
                
                for ans in prior_answers:
                    await repo_create("assessment_answer", {
                        "session_id": new_sess_id,
                        "question_id": ans["question_id"],
                        "answer": ans["answer"],
                        "comments": ans.get("comments") or "",
                        "evidence_url": ans.get("evidence_url") or "",
                        "updated_at": datetime.now(timezone.utc).isoformat() + "Z"
                    })
                    
        # Emit activity for new session
        assess_record = assess[0]
        cust_id = assess_record.get("customer_id")
        if cust_id:
            from api.routers.activity_emitter import emit_activity
            await emit_activity(
                customer_id=str(cust_id),
                activity_type="assessment_started",
                description=f"Audit session \"{data.session_name}\" started",
                metadata={"session_id": new_sess_id, "assessment_id": assess_id},
            )

        return AssessmentSessionResponse(
            id=new_sess_id,
            assessment_id=str(new_sess["assessment_id"]),
            session_name=str(new_sess["session_name"]),
            created_at=str(new_sess["created_at"]),
            completed_at=new_sess.get("completed_at"),
            status=str(new_sess["status"]),
            version_lock=new_sess.get("version_lock")
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating audit session: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/assessments/{assessment_id}/sessions", response_model=List[AssessmentSessionResponse])
async def get_sessions(assessment_id: str):
    """List all audit sessions under a specific assessment."""
    try:
        assess_id = assessment_id
        if ":" not in assess_id:
            assess_id = f"assessment:{assessment_id}"
            
        results = await repo_query(
            "SELECT * FROM assessment_session WHERE type::string(assessment_id) = type::string($assess_id) ORDER BY created_at DESC",
            {"assess_id": assess_id}
        )
        
        return [
            AssessmentSessionResponse(
                id=str(row["id"]),
                assessment_id=str(row["assessment_id"]),
                session_name=str(row["session_name"]),
                created_at=str(row["created_at"]),
                completed_at=row.get("completed_at"),
                status=str(row["status"]),
                version_lock=row.get("version_lock")
            )
            for row in results
        ]
    except Exception as e:
        logger.error(f"Error fetching sessions: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sessions/{session_id}/questions")
async def get_session_questions(session_id: str):
    """Retrieve all framework questions hydrated with existing session answers in a single joined payload."""
    try:
        sess_id = session_id
        if ":" not in sess_id:
            sess_id = f"assessment_session:{session_id}"
            
        sess_check = await repo_query("SELECT * FROM $id", {"id": ensure_record_id(sess_id)})
        if not sess_check:
            raise HTTPException(status_code=404, detail="Audit session not found")
            
        session = sess_check[0]
        fw_id = session.get("version_lock")
        
        # Resolve active prefixes for Components framework linkage
        active_prefixes = await get_active_cset_prefixes_for_session(sess_id, fw_id)

        # Load all standard questions
        questions = await repo_query(
            "SELECT * FROM question WHERE regulation_id = $fw_id ORDER BY standard_code ASC",
            {"fw_id": fw_id}
        )
        
        # Load all answered responses in this session
        answers = await repo_query(
            "SELECT * FROM assessment_answer WHERE type::string(session_id) = type::string($session_id)",
            {"session_id": sess_id}
        )
        
        # Create quick lookup map for answered questions
        answer_map = {ans["question_id"]: ans for ans in answers}
        
        results = []
        for q in questions:
            q_id = str(q["id"])
            existing_ans = answer_map.get(q_id)
            
            ans_val = "U"
            if existing_ans:
                ans_val = existing_ans["answer"]
                
            # Under "no exclusion" policy, mark questions as NA if their component type is not present on canvas
            if active_prefixes is not None:
                std_code = q.get("standard_code") or ""
                is_active = False
                for prefix in active_prefixes:
                    if std_code.strip().startswith(prefix):
                        is_active = True
                        break
                if not is_active:
                    ans_val = "NA"
            
            results.append({
                "question_id": q_id,
                "standard_code": q.get("standard_code") or "",
                "question_text": q.get("question_text") or "",
                "description": q.get("description") or "",
                "purdue_level": int(q.get("purdue_level") or 0),
                "category": q.get("category") or "Control",
                "answer": ans_val,
                "comments": existing_ans["comments"] if existing_ans else "",
                "evidence_url": existing_ans["evidence_url"] if existing_ans else "",
                "updated_at": existing_ans["updated_at"] if existing_ans else ""
            })
            
        return results
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching session questions: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/sessions/{session_id}/answers/{question_id}", response_model=AssessmentAnswerResponse)
async def update_answer(session_id: str, question_id: str, data: AssessmentAnswerUpdate):
    """Autosave or update an individual compliance answer inside an audit session."""
    try:
        sess_id = session_id
        if ":" not in sess_id:
            sess_id = f"assessment_session:{session_id}"
            
        q_id = question_id
        if ":" not in q_id:
            q_id = f"question:{question_id}"
            
        # Verify session exists and is active
        sess_check = await repo_query("SELECT * FROM $id", {"id": ensure_record_id(sess_id)})
        if not sess_check:
            raise HTTPException(status_code=404, detail="Audit session not found")
            
        if sess_check[0].get("status") == "COMPLETED":
            raise HTTPException(status_code=400, detail="Cannot edit answers inside a completed audit session")
            
        # Check if answer record already exists
        existing = await repo_query(
            "SELECT id FROM assessment_answer WHERE type::string(session_id) = type::string($session_id) AND type::string(question_id) = type::string($question_id)",
            {"session_id": sess_id, "question_id": q_id}
        )
        
        now_str = datetime.now(timezone.utc).isoformat() + "Z"
        
        if existing:
            ans_record_id = str(existing[0]["id"])
            result = await repo_update("assessment_answer", ans_record_id, {
                "answer": data.answer,
                "comments": data.comments or "",
                "evidence_url": data.evidence_url or "",
                "updated_at": now_str
            })
        else:
            result = await repo_create("assessment_answer", {
                "session_id": sess_id,
                "question_id": q_id,
                "answer": data.answer,
                "comments": data.comments or "",
                "evidence_url": data.evidence_url or "",
                "updated_at": now_str
            })
            
        if isinstance(result, list):
            result = result[0]
            
        return AssessmentAnswerResponse(
            id=str(result["id"]),
            session_id=str(result["session_id"]),
            question_id=str(result["question_id"]),
            answer=str(result["answer"]),
            comments=str(result.get("comments") or ""),
            evidence_url=str(result.get("evidence_url") or ""),
            updated_at=str(result["updated_at"])
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error saving wizard answer: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/sessions/{session_id}/complete", response_model=AssessmentSessionResponse)
async def complete_session(session_id: str):
    """Finalize and lock an active audit session milestone and store a static compliance snapshot."""
    try:
        sess_id = session_id
        if ":" not in sess_id:
            sess_id = f"assessment_session:{session_id}"
            
        sess_check = await repo_query("SELECT * FROM $id", {"id": ensure_record_id(sess_id)})
        if not sess_check:
            raise HTTPException(status_code=404, detail="Audit session not found")
            
        session_record = sess_check[0]
        fw_id = session_record.get("version_lock")

        # Load standard questions & custom fields
        questions = await repo_query(
            "SELECT * FROM question WHERE regulation_id = $fw_id",
            {"fw_id": fw_id}
        )
        
        # Load answers
        answers = await repo_query(
            "SELECT * FROM assessment_answer WHERE type::string(session_id) = type::string($session_id)",
            {"session_id": sess_id}
        )
        
        answer_map = {ans["question_id"]: ans for ans in answers}
        
        # Resolve active prefixes for Components framework linkage
        active_prefixes = await get_active_cset_prefixes_for_session(sess_id, fw_id)
        
        total_questions = len(questions)
        
        yes_count = 0
        no_count = 0
        na_count = 0
        alt_count = 0
        answered_count = 0
        
        # Calculate statistics
        for q in questions:
            q_id = str(q["id"])
            ans = answer_map.get(q_id)
            
            ans_val = "U"
            if ans:
                ans_val = ans["answer"]
                
            # Enforce "no exclusion" asset register linkage
            is_active = True
            if active_prefixes is not None:
                std_code = q.get("standard_code") or ""
                is_active = False
                for prefix in active_prefixes:
                    if std_code.strip().startswith(prefix):
                        is_active = True
                        break
                if not is_active:
                    ans_val = "NA"
            
            if ans_val == "Y":
                yes_count += 1
            elif ans_val == "N":
                no_count += 1
            elif ans_val == "NA":
                na_count += 1
            elif ans_val == "ALT":
                alt_count += 1
            elif ans_val == "U":
                no_count += 1  # Unanswered defaults to NO
                
            if is_active and ans and ans["answer"] != "U":
                answered_count += 1
        
        denominator = total_questions - na_count
        compliance_score = ((yes_count + alt_count) / denominator) * 100 if denominator > 0 else 0.0
        
        # Category coverage analysis
        category_map = {}
        for q in questions:
            cat = q.get("category") or "Control"
            if cat not in category_map:
                category_map[cat] = {"total": 0, "answered": 0, "yes_count": 0}
            
            q_id = str(q["id"])
            ans = answer_map.get(q_id)
            
            ans_val = "U"
            if ans:
                ans_val = ans["answer"]
                
            is_active = True
            if active_prefixes is not None:
                std_code = q.get("standard_code") or ""
                is_active = False
                for prefix in active_prefixes:
                    if std_code.strip().startswith(prefix):
                        is_active = True
                        break
                if not is_active:
                    ans_val = "NA"
            
            if ans_val != "NA":
                category_map[cat]["total"] += 1
                if ans and ans["answer"] != "U":
                    category_map[cat]["answered"] += 1
                if ans_val in ["Y", "ALT"]:
                    category_map[cat]["yes_count"] += 1
                    
        category_coverage = []
        for cat, metrics in category_map.items():
            tot = metrics["total"]
            yes = metrics["yes_count"]
            score = (yes / tot) * 100 if tot > 0 else 0.0
            category_coverage.append({
                "category": cat,
                "total": tot,
                "answered": metrics["answered"],
                "yes_count": yes,
                "score": score
            })

        compliance_snapshot = {
            "compliance_score": compliance_score,
            "total_questions": total_questions,
            "answered_count": answered_count,
            "yes_count": yes_count,
            "no_count": no_count,
            "na_count": na_count,
            "alt_count": alt_count,
            "category_coverage": category_coverage
        }

        result = await repo_update("assessment_session", sess_id, {
            "status": "COMPLETED",
            "completed_at": datetime.now(timezone.utc).isoformat() + "Z",
            "compliance_snapshot": compliance_snapshot
        })
        
        if isinstance(result, list):
            result = result[0]
            
        return AssessmentSessionResponse(
            id=str(result["id"]),
            assessment_id=str(result["assessment_id"]),
            session_name=str(result["session_name"]),
            created_at=str(result["created_at"]),
            completed_at=result.get("completed_at"),
            status=str(result["status"]),
            version_lock=result.get("version_lock"),
            compliance_snapshot=result.get("compliance_snapshot")
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error locking session: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sessions/{session_id}/report", response_model=AssessmentReportResponse)
async def get_session_report(session_id: str):
    """Generate CSET-style auditing report: scoring indices, category coverages, and prioritizations."""
    try:
        sess_id = session_id
        if ":" not in sess_id:
            sess_id = f"assessment_session:{session_id}"
            
        sess_check = await repo_query("SELECT * FROM $id", {"id": ensure_record_id(sess_id)})
        if not sess_check:
            raise HTTPException(status_code=404, detail="Audit session not found")
            
        session = sess_check[0]
        fw_id = session.get("version_lock")
        
        # Load standard questions & custom fields
        questions = await repo_query(
            "SELECT * FROM question WHERE regulation_id = $fw_id",
            {"fw_id": fw_id}
        )
        
        # Load answers
        answers = await repo_query(
            "SELECT * FROM assessment_answer WHERE type::string(session_id) = type::string($session_id)",
            {"session_id": sess_id}
        )
        
        answer_map = {ans["question_id"]: ans for ans in answers}
        
        # Resolve active prefixes for Components framework linkage
        active_prefixes = await get_active_cset_prefixes_for_session(sess_id, fw_id)
        
        total_questions = len(questions)
        
        yes_count = 0
        no_count = 0
        na_count = 0
        alt_count = 0
        answered_count = 0
        
        # Calculate statistics
        for q in questions:
            q_id = str(q["id"])
            ans = answer_map.get(q_id)
            
            ans_val = "U"
            if ans:
                ans_val = ans["answer"]
                
            # Enforce "no exclusion" asset register linkage
            is_active = True
            if active_prefixes is not None:
                std_code = q.get("standard_code") or ""
                is_active = False
                for prefix in active_prefixes:
                    if std_code.strip().startswith(prefix):
                        is_active = True
                        break
                if not is_active:
                    ans_val = "NA"
            
            if ans_val == "Y":
                yes_count += 1
            elif ans_val == "N":
                no_count += 1
            elif ans_val == "NA":
                na_count += 1
            elif ans_val == "ALT":
                alt_count += 1
            elif ans_val == "U":
                no_count += 1  # Unanswered defaults to NO
                
            if is_active and ans and ans["answer"] != "U":
                answered_count += 1
        
        completion_percentage = (answered_count / total_questions) * 100 if total_questions > 0 else 0.0
        
        # Compliance score: YES + ALT compared against all non-N/A questions
        denominator = total_questions - na_count
        compliance_score = ((yes_count + alt_count) / denominator) * 100 if denominator > 0 else 0.0
        
        # Category coverage analysis
        category_map = {}
        for q in questions:
            cat = q.get("category") or "Control"
            if cat not in category_map:
                category_map[cat] = {"total": 0, "answered": 0, "yes_count": 0}
            
            q_id = str(q["id"])
            ans = answer_map.get(q_id)
            
            ans_val = "U"
            if ans:
                ans_val = ans["answer"]
                
            is_active = True
            if active_prefixes is not None:
                std_code = q.get("standard_code") or ""
                is_active = False
                for prefix in active_prefixes:
                    if std_code.strip().startswith(prefix):
                        is_active = True
                        break
                if not is_active:
                    ans_val = "NA"
            
            if ans_val != "NA":
                category_map[cat]["total"] += 1
                if ans and ans["answer"] != "U":
                    category_map[cat]["answered"] += 1
                if ans_val in ["Y", "ALT"]:
                    category_map[cat]["yes_count"] += 1
                    
        category_coverage = []
        for cat, metrics in category_map.items():
            tot = metrics["total"]
            yes = metrics["yes_count"]
            score = (yes / tot) * 100 if tot > 0 else 0.0
            category_coverage.append(
                CategoryCoverage(
                    category=cat,
                    total=tot,
                    answered=metrics["answered"],
                    yes_count=yes,
                    score=score
                )
            )
            
        # Prioritized recommendations
        prioritized_recommendations = []
        for q in questions:
            q_id = str(q["id"])
            ans = answer_map.get(q_id)
            
            ans_val = "U"
            if ans:
                ans_val = ans["answer"]
                
            is_active = True
            if active_prefixes is not None:
                std_code = q.get("standard_code") or ""
                is_active = False
                for prefix in active_prefixes:
                    if std_code.strip().startswith(prefix):
                        is_active = True
                        break
                if not is_active:
                    ans_val = "NA"
            
            # Recommend fixing if active and (unanswered or explicitly answered 'NO')
            if is_active and (ans_val == "U" or ans_val == "N"):
                p_level = int(q.get("purdue_level") or 0)
                
                # Priority mapping: Purdue Level 1 & 2 is Critical (kinetic boundary)
                if p_level in [1, 2]:
                    priority = "Critical"
                    order = 1
                elif p_level == 3:
                    priority = "High"
                    order = 2
                else:
                    priority = "Medium"
                    order = 3
                    
                prioritized_recommendations.append({
                    "question_id": q_id,
                    "standard_code": q.get("standard_code") or "",
                    "question_text": q.get("question_text") or "",
                    "category": q.get("category") or "Control",
                    "purdue_level": p_level,
                    "description": q.get("description") or "",
                    "priority": priority,
                    "order": order
                })
                
        # Sort recommendations by order first, then standard code
        prioritized_recommendations.sort(key=lambda x: (x["order"], x["standard_code"]))
        
        stats = AssessmentReportStats(
            total_questions=total_questions,
            answered_count=answered_count,
            yes_count=yes_count,
            no_count=no_count,
            na_count=na_count,
            alt_count=alt_count,
            completion_percentage=completion_percentage,
            compliance_score=compliance_score
        )
        
        return AssessmentReportResponse(
            session_id=sess_id,
            session_name=str(session["session_name"]),
            framework_id=str(fw_id),
            stats=stats,
            category_coverage=category_coverage,
            prioritized_recommendations=prioritized_recommendations
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error compiling session report: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/assessments/{assessment_id}/trends")
async def get_assessment_trends(assessment_id: str):
    """Retrieve chronological trends and deltas across all completed sessions in an assessment."""
    try:
        assess_id = assessment_id
        if ":" not in assess_id:
            assess_id = f"assessment:{assessment_id}"
            
        sessions = await repo_query(
            "SELECT * FROM assessment_session WHERE type::string(assessment_id) = type::string($assess_id) ORDER BY created_at ASC",
            {"assess_id": assess_id}
        )
        
        trends = []
        previous_score = None
        
        for sess in sessions:
            sess_id = str(sess["id"])
            fw_id = sess.get("version_lock")
            
            # Load answers & questions for scoring
            q_count_res = await repo_query(
                "SELECT count() FROM question WHERE regulation_id = $fw_id GROUP ALL",
                {"fw_id": fw_id}
            )
            total_questions = q_count_res[0]["count"] if q_count_res else 0
            
            ans_res = await repo_query(
                "SELECT answer FROM assessment_answer WHERE type::string(session_id) = type::string($session_id)",
                {"session_id": sess_id}
            )
            
            yes_count = sum(1 for a in ans_res if a["answer"] == "Y")
            alt_count = sum(1 for a in ans_res if a["answer"] == "ALT")
            na_count = sum(1 for a in ans_res if a["answer"] == "NA")
            
            denominator = total_questions - na_count
            compliance_score = ((yes_count + alt_count) / denominator) * 100 if denominator > 0 else 0.0
            
            delta = 0.0
            if previous_score is not None:
                delta = compliance_score - previous_score
                
            trends.append({
                "session_id": sess_id,
                "session_name": str(sess["session_name"]),
                "created_at": str(sess["created_at"]),
                "completed_at": sess.get("completed_at"),
                "status": str(sess["status"]),
                "compliance_score": compliance_score,
                "delta": delta
            })
            
            if sess["status"] == "COMPLETED":
                previous_score = compliance_score
                
        return trends
    except Exception as e:
        logger.error(f"Error computing auditing trends: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


def _build_session_csv_stream(report: AssessmentReportResponse):
    import csv
    import io
    buf = io.StringIO()
    headers = ["Priority", "Standard Code", "Category", "Purdue Level", "Question Text", "Guidance Description"]
    writer = csv.writer(buf)
    writer.writerow(headers)
    for rec in report.prioritized_recommendations:
        writer.writerow([
            rec.get("priority", "Medium"),
            rec.get("standard_code", ""),
            rec.get("category", ""),
            rec.get("purdue_level", 0),
            rec.get("question_text", ""),
            rec.get("description", "")
        ])
    buf.seek(0)
    return buf


def _build_session_xlsx_bytes(report: AssessmentReportResponse) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    
    # 1. Summary Sheet
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    # Fonts & Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=13, bold=True, color="1F497D")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    accent_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Title
    ws1["A1"] = f"Compliance Audit Summary: {report.session_name}"
    ws1["A1"].font = title_font
    
    ws1["A2"] = f"Framework: {report.framework_id.replace('regulation:', '')} | Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
    ws1["A2"].font = Font(name="Calibri", size=10, italic=True)
    
    # Score Callout
    ws1["D4"] = "Compliance Rating"
    ws1["D4"].font = bold_font
    ws1["D5"] = f"{report.stats.compliance_score:.1f}%"
    ws1["D5"].font = Font(name="Calibri", size=24, bold=True, color="2E75B6")
    ws1["D5"].fill = accent_fill
    ws1["D5"].alignment = Alignment(horizontal="center")
    
    # Stats Overview table
    ws1["A4"] = "AUDITING METRICS"
    ws1["A4"].font = section_font
    
    metrics = [
        ("Total Questions", report.stats.total_questions),
        ("Completed Answers", report.stats.answered_count),
        ("Yes Compliances", report.stats.yes_count),
        ("Alternative (ALT) Controls", report.stats.alt_count),
        ("Identified Gaps (No/Unanswered)", report.stats.no_count),
        ("N/A (Exclusions)", report.stats.na_count),
        ("Completion Progress", f"{report.stats.completion_percentage:.1f}%")
    ]
    
    ws1.cell(row=5, column=1, value="Metric").font = header_font
    ws1.cell(row=5, column=1).fill = header_fill
    ws1.cell(row=5, column=2, value="Value").font = header_font
    ws1.cell(row=5, column=2).fill = header_fill
    
    for idx, (metric, val) in enumerate(metrics, start=6):
        c1 = ws1.cell(row=idx, column=1, value=metric)
        c2 = ws1.cell(row=idx, column=2, value=val)
        c1.font = regular_font
        c2.font = bold_font
        c1.border = thin_border
        c2.border = thin_border
        
    # Category Coverage Table
    start_row = 15
    ws1.cell(row=start_row, column=1, value="CATEGORY COVERAGE INDEX").font = section_font
    
    cat_headers = ["Category Family", "Total", "Answered", "Yes/ALT", "Score (%)"]
    for col_idx, text in enumerate(cat_headers, start=1):
        cell = ws1.cell(row=start_row + 1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        
    for row_idx, cat in enumerate(report.category_coverage, start=start_row + 2):
        r_fill = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
        
        c1 = ws1.cell(row=row_idx, column=1, value=cat.category)
        c2 = ws1.cell(row=row_idx, column=2, value=cat.total)
        c3 = ws1.cell(row=row_idx, column=3, value=cat.answered)
        c4 = ws1.cell(row=row_idx, column=4, value=cat.yes_count)
        c5 = ws1.cell(row=row_idx, column=5, value=f"{cat.score:.1f}%")
        
        for cell in [c1, c2, c3, c4, c5]:
            cell.font = regular_font
            cell.border = thin_border
            if r_fill.fill_type:
                cell.fill = r_fill
        c5.font = bold_font
        
    # Set Column Widths for Summary
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # 2. Roadmap Sheet
    ws2 = wb.create_sheet(title="Remediation Roadmap")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = "Prioritized Remediation Roadmap (Gap Fixes)"
    ws2["A1"].font = section_font
    
    roadmap_headers = ["Priority", "Standard Code", "Requirement Directive Description", "Category Family", "Purdue Level", "Technical Guidance Details"]
    for col_idx, text in enumerate(roadmap_headers, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        
    priority_colors = {
        "Critical": "FCE4D6", # Soft red
        "High": "FFF2CC",     # Soft orange
        "Medium": "F2F2F2"    # Light gray
    }
    
    for row_idx, rec in enumerate(report.prioritized_recommendations, start=4):
        prio = rec.get("priority", "Medium")
        prio_fill = PatternFill(start_color=priority_colors.get(prio, "FFFFFF"), end_color=priority_colors.get(prio, "FFFFFF"), fill_type="solid")
        
        c1 = ws2.cell(row=row_idx, column=1, value=prio)
        c2 = ws2.cell(row=row_idx, column=2, value=rec.get("standard_code", ""))
        c3 = ws2.cell(row=row_idx, column=3, value=rec.get("question_text", ""))
        c4 = ws2.cell(row=row_idx, column=4, value=rec.get("category", ""))
        c5 = ws2.cell(row=row_idx, column=5, value=rec.get("purdue_level", 0))
        c6 = ws2.cell(row=row_idx, column=6, value=rec.get("description", ""))
        
        for cell in [c1, c2, c3, c4, c5, c6]:
            cell.font = regular_font
            cell.border = thin_border
            
        c1.fill = prio_fill
        c1.alignment = Alignment(horizontal="center")
        c2.font = bold_font
        c5.alignment = Alignment(horizontal="center")
        
    # Wrap text and set dimensions for Roadmap sheet
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 50
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 50
    
    for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
    for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/sessions/{session_id}/diff/{compare_session_id}", response_model=AssessmentSessionDiffResponse)
async def get_session_diff(session_id: str, compare_session_id: str):
    """Retrieve question-level differences (rating, comments, evidence) between two sessions."""
    try:
        base_sess_id = session_id
        if ":" not in base_sess_id:
            base_sess_id = f"assessment_session:{session_id}"
        compare_sess_id = compare_session_id
        if ":" not in compare_sess_id:
            compare_sess_id = f"assessment_session:{compare_session_id}"
            
        sess_check_base = await repo_query("SELECT * FROM $id", {"id": ensure_record_id(base_sess_id)})
        if not sess_check_base:
            raise HTTPException(status_code=404, detail="Base audit session not found")
            
        sess_check_compare = await repo_query("SELECT * FROM $id", {"id": ensure_record_id(compare_sess_id)})
        if not sess_check_compare:
            raise HTTPException(status_code=404, detail="Comparison audit session not found")
            
        base_sess = sess_check_base[0]
        base_fw = base_sess.get("version_lock")
        
        compare_sess = sess_check_compare[0]
        compare_fw = compare_sess.get("version_lock")
        
        # Load questions
        base_questions = await repo_query(
            "SELECT * FROM question WHERE regulation_id = $fw_id",
            {"fw_id": base_fw}
        )
        if base_fw == compare_fw:
            compare_questions = base_questions
        else:
            compare_questions = await repo_query(
                "SELECT * FROM question WHERE regulation_id = $fw_id",
                {"fw_id": compare_fw}
            )
            
        # Load answers
        base_answers = await repo_query(
            "SELECT * FROM assessment_answer WHERE type::string(session_id) = type::string($session_id)",
            {"session_id": base_sess_id}
        )
        compare_answers = await repo_query(
            "SELECT * FROM assessment_answer WHERE type::string(session_id) = type::string($session_id)",
            {"session_id": compare_sess_id}
        )
        
        base_ans_map = {str(ans["question_id"]): ans for ans in base_answers}
        compare_ans_map = {str(ans["question_id"]): ans for ans in compare_answers}
        
        # Maps for questions by standard code (primary key) and id (fallback)
        base_q_by_code = {q.get("standard_code"): q for q in base_questions if q.get("standard_code")}
        base_q_by_id = {str(q["id"]): q for q in base_questions}
        
        compare_q_by_code = {q.get("standard_code"): q for q in compare_questions if q.get("standard_code")}
        compare_q_by_id = {str(q["id"]): q for q in compare_questions}
        
        # Gather all keys to iterate over
        all_keys = set()
        for q in base_questions:
            all_keys.add(q.get("standard_code") or str(q["id"]))
        for q in compare_questions:
            all_keys.add(q.get("standard_code") or str(q["id"]))
            
        active_prefixes_base = await get_active_cset_prefixes_for_session(base_sess_id, base_fw)
        active_prefixes_compare = await get_active_cset_prefixes_for_session(compare_sess_id, compare_fw)
        
        differences = []
        for key in sorted(all_keys):
            q_base = base_q_by_code.get(key) or base_q_by_id.get(key)
            q_compare = compare_q_by_code.get(key) or compare_q_by_id.get(key)
            
            q = q_compare or q_base
            if not q:
                continue
                
            q_id = str(q["id"])
            standard_code = q.get("standard_code") or ""
            question_text = q.get("question_text") or ""
            category = q.get("category") or "Control"
            purdue_level = int(q.get("purdue_level") or 0)
            
            # Base answer mapping
            base_ans_val = "U"
            base_comments = ""
            base_evidence = ""
            if q_base:
                ans_b_obj = base_ans_map.get(str(q_base["id"]))
                if ans_b_obj:
                    base_ans_val = ans_b_obj["answer"]
                    base_comments = ans_b_obj.get("comments") or ""
                    base_evidence = ans_b_obj.get("evidence_url") or ""
                    
                if active_prefixes_base is not None:
                    std_c_b = q_base.get("standard_code") or ""
                    is_act = False
                    for prefix in active_prefixes_base:
                        if std_c_b.strip().startswith(prefix):
                            is_act = True
                            break
                    if not is_act:
                        base_ans_val = "NA"
                        
            # Compare answer mapping
            compare_ans_val = "U"
            compare_comments = ""
            compare_evidence = ""
            if q_compare:
                ans_c_obj = compare_ans_map.get(str(q_compare["id"]))
                if ans_c_obj:
                    compare_ans_val = ans_c_obj["answer"]
                    compare_comments = ans_c_obj.get("comments") or ""
                    compare_evidence = ans_c_obj.get("evidence_url") or ""
                    
                if active_prefixes_compare is not None:
                    std_c_c = q_compare.get("standard_code") or ""
                    is_act = False
                    for prefix in active_prefixes_compare:
                        if std_c_c.strip().startswith(prefix):
                            is_act = True
                            break
                    if not is_act:
                        compare_ans_val = "NA"
                        
            has_changed = (
                (base_ans_val != compare_ans_val) or
                (base_comments != compare_comments) or
                (base_evidence != compare_evidence)
            )
            
            differences.append(AssessmentSessionDiffItem(
                question_id=q_id,
                standard_code=standard_code,
                question_text=question_text,
                category=category,
                purdue_level=purdue_level,
                base_answer=base_ans_val,
                compare_answer=compare_ans_val,
                base_comments=base_comments,
                compare_comments=compare_comments,
                base_evidence=base_evidence,
                compare_evidence=compare_evidence,
                has_changed=has_changed
            ))
            
        return AssessmentSessionDiffResponse(
            base_session_id=base_sess_id,
            base_session_name=str(base_sess["session_name"]),
            compare_session_id=compare_sess_id,
            compare_session_name=str(compare_sess["session_name"]),
            differences=differences
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error computing session diff: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sessions/{session_id}/export")
async def export_session_report(
    session_id: str,
    format: str = Query("xlsx", description="Export format: csv or xlsx")
):
    """Export the compliance report metrics and prioritized gaps remediation list."""
    from fastapi.responses import StreamingResponse
    import io
    try:
        report = await get_session_report(session_id)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        safe_name = "".join(c for c in report.session_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
        
        if format.lower() == "xlsx":
            xlsx_data = _build_session_xlsx_bytes(report)
            return StreamingResponse(
                io.BytesIO(xlsx_data),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="compliance_report_{safe_name}_{timestamp}.xlsx"'},
            )
        else:
            csv_buf = _build_session_csv_stream(report)
            return StreamingResponse(
                iter([csv_buf.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="compliance_report_{safe_name}_{timestamp}.csv"'},
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting session report: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


DB_TO_FRONTEND_MAP = {
    'Cfats': 'CFATS_RBPS',
    'CPG': 'CISA_CPG',
    'SP800_82_V3': 'NIST_800_82',
    'SP800_82_V2': 'NIST_800_82',
    'C800_53_R5_V2': 'NIST_800_53',
    'C800_53_R4_71': 'NIST_800_53',
    'NCSF_V2': 'NIST_CSF',
    'NCSF_V1': 'NIST_CSF',
    'CSC_V8': 'CIS_CONTROLS',
    'Cnssi_1253': 'CNSSI_1253',
    'AWWA': 'AWWA_G430',
    'TSA2018': 'TSA_RAIL',
    'Tsa': 'TSA_PIPELINE',
    'COBIT_2019': 'COBIT_2019',
    'SOC_2': 'SOC_2',
    'ISA_62443': 'IEC_62443_3_3',
    'CMMC': 'CMMC_L1',
    'Universal': 'NIS2',
}


@router.get("/customers/{customer_id}/compliance-rollup", response_model=CustomerComplianceRollup)
async def get_customer_compliance_rollup(customer_id: str):
    """Generate isolated facility-level audit statistics and combined rollup reports for an organization."""
    try:
        cust_id = customer_id
        if ":" not in cust_id:
            cust_id = f"customer:{cust_id}"

        # 1. Fetch customer details to verify existence
        cust_check = await repo_query("SELECT id FROM $id", {"id": ensure_record_id(cust_id)})
        if not cust_check:
            raise HTTPException(status_code=404, detail="Customer profile not found")

        # 2. Fetch all locations for the customer
        locations = await repo_query(
            "SELECT id, facility_name, facility_type FROM location WHERE type::string(customer_id) = type::string($cust_id)",
            {"cust_id": cust_id}
        )
        location_map = {str(loc["id"]): f"{loc['facility_name']} ({loc['facility_type']})" if loc.get("facility_type") else str(loc["facility_name"]) for loc in locations}

        # 3. Fetch all regulations/frameworks names
        regulations = await repo_query("SELECT id, name FROM regulation")
        reg_map = {str(reg["id"]): str(reg.get("name") or "") for reg in regulations}

        # 4. Fetch all assessments for the customer
        assessments = await repo_query(
            "SELECT * FROM assessment WHERE type::string(customer_id) = type::string($cust_id) ORDER BY created_at DESC",
            {"cust_id": cust_id}
        )

        # 5. Build framework rollups
        framework_data = {}

        for assess in assessments:
            assess_id = str(assess["id"])
            db_fw_id = str(assess["framework_id"])
            raw_fw_id = db_fw_id.replace("regulation:", "")
            frontend_fw_id = DB_TO_FRONTEND_MAP.get(raw_fw_id, raw_fw_id)
            fw_name = reg_map.get(db_fw_id, raw_fw_id)

            # Retrieve sessions for this assessment
            sessions = await repo_query(
                "SELECT * FROM assessment_session WHERE type::string(assessment_id) = type::string($assess_id) ORDER BY created_at DESC",
                {"assess_id": assess_id}
            )

            # Determine facility info
            loc_id = assess.get("location_id")
            loc_str = str(loc_id) if loc_id else None
            facility_name = location_map.get(loc_str, "Organization-Wide") if loc_str else "Organization-Wide"

            facility_rollup = FacilityRollup(
                location_id=loc_str,
                facility_name=facility_name,
                status="NOT_STARTED",
                completion_percentage=0.0,
                compliance_score=0.0
            )

            if sessions:
                latest_session = sessions[0]
                sess_id = str(latest_session["id"])
                
                # Fetch question count
                q_count_res = await repo_query(
                    "SELECT count() FROM question WHERE regulation_id = $fw_id GROUP ALL",
                    {"fw_id": db_fw_id}
                )
                total_questions = q_count_res[0]["count"] if q_count_res else 0

                if latest_session.get("status") == "COMPLETED" and latest_session.get("compliance_snapshot"):
                    snapshot = latest_session["compliance_snapshot"]
                    completion_percentage = 100.0
                    compliance_score = float(snapshot.get("compliance_score") or 0.0)
                else:
                    # In-progress calculation
                    answers = await repo_query(
                        "SELECT answer FROM assessment_answer WHERE type::string(session_id) = type::string($session_id)",
                        {"session_id": sess_id}
                    )
                    
                    yes_count = sum(1 for a in answers if a["answer"] == "Y")
                    alt_count = sum(1 for a in answers if a["answer"] == "ALT")
                    na_count = sum(1 for a in answers if a["answer"] == "NA")
                    answered_count = sum(1 for a in answers if a["answer"] != "U")

                    completion_percentage = (answered_count / total_questions) * 100 if total_questions > 0 else 0.0
                    denominator = total_questions - na_count
                    compliance_score = ((yes_count + alt_count) / denominator) * 100 if denominator > 0 else 0.0

                facility_rollup.session_id = sess_id
                facility_rollup.session_name = str(latest_session["session_name"])
                facility_rollup.status = str(latest_session["status"])
                facility_rollup.completion_percentage = completion_percentage
                facility_rollup.compliance_score = compliance_score
                facility_rollup.last_updated = str(latest_session.get("completed_at") or latest_session.get("created_at") or "")

            if frontend_fw_id not in framework_data:
                framework_data[frontend_fw_id] = {
                    "framework_id": frontend_fw_id,
                    "framework_name": fw_name,
                    "facilities": []
                }
            
            framework_data[frontend_fw_id]["facilities"].append(facility_rollup)

        # 6. Compute framework averages & compile rollups
        framework_rollups = []
        for fw_id, fw_info in framework_data.items():
            facs = fw_info["facilities"]
            total_facs = len(facs)
            
            # Average calculations across initialized facilities
            avg_score = sum(f.compliance_score for f in facs) / total_facs if total_facs > 0 else 0.0
            avg_completion = sum(f.completion_percentage for f in facs) / total_facs if total_facs > 0 else 0.0

            framework_rollups.append(FrameworkRollup(
                framework_id=fw_id,
                framework_name=fw_info["framework_name"],
                facilities=facs,
                average_compliance_score=avg_score,
                average_completion_percentage=avg_completion,
                total_facilities_assessed=total_facs
            ))

        return CustomerComplianceRollup(
            customer_id=cust_id,
            frameworks=framework_rollups
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error computing compliance rollup: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
